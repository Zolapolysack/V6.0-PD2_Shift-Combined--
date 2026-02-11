#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ตรวจสอบไฟล์ Excel รายงานสินค้าสำเร็จรูป
"""

import openpyxl

# Load Excel file
wb = openpyxl.load_workbook('รายงานสินค้าสำเร็จรูป.xlsx')

print("="*80)
print("ตรวจสอบไฟล์ Excel: รายงานสินค้าสำเร็จรูป.xlsx")
print("="*80)

print(f"\nSheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*80}")
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")
    
    print(f"\nFirst 10 rows:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        print(f"Row {i}: {row}")
    
    # Check for merged cells
    if ws.merged_cells:
        print(f"\nMerged Cells: {len(ws.merged_cells.ranges)} ranges")
        for merged_range in list(ws.merged_cells.ranges)[:5]:
            print(f"  - {merged_range}")
    
    # Check for images
    if hasattr(ws, '_images') and ws._images:
        print(f"\nImages: {len(ws._images)}")

print("\n" + "="*80)
print("เสร็จสิ้นการตรวจสอบ")
print("="*80)
