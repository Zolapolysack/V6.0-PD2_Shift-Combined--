#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
วิเคราะห์ Template อย่างละเอียด 100%
เก็บทุกรายละเอียด: Font, Size, Color, Border, Alignment, Images, etc.
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
import json

def analyze_cell_style(cell):
    """วิเคราะห์ style ของ cell อย่างละเอียด"""
    
    def rgb_to_str(rgb_obj):
        """แปลง RGB object เป็น string"""
        if rgb_obj and hasattr(rgb_obj, 'rgb'):
            return rgb_obj.rgb
        return None
    
    style = {
        'value': str(cell.value) if cell.value else None,
        'data_type': cell.data_type,
        'font': {
            'name': cell.font.name,
            'size': cell.font.size,
            'bold': cell.font.bold,
            'italic': cell.font.italic,
            'underline': cell.font.underline,
            'strike': cell.font.strike,
            'color': rgb_to_str(cell.font.color),
        },
        'fill': {
            'fgColor': rgb_to_str(cell.fill.fgColor) if cell.fill else None,
            'bgColor': rgb_to_str(cell.fill.bgColor) if cell.fill else None,
            'patternType': cell.fill.patternType if cell.fill else None,
        },
        'border': {
            'left': {
                'style': cell.border.left.style if cell.border and cell.border.left else None,
                'color': rgb_to_str(cell.border.left.color) if cell.border and cell.border.left else None,
            },
            'right': {
                'style': cell.border.right.style if cell.border and cell.border.right else None,
                'color': rgb_to_str(cell.border.right.color) if cell.border and cell.border.right else None,
            },
            'top': {
                'style': cell.border.top.style if cell.border and cell.border.top else None,
                'color': rgb_to_str(cell.border.top.color) if cell.border and cell.border.top else None,
            },
            'bottom': {
                'style': cell.border.bottom.style if cell.border and cell.border.bottom else None,
                'color': rgb_to_str(cell.border.bottom.color) if cell.border and cell.border.bottom else None,
            },
        },
        'alignment': {
            'horizontal': cell.alignment.horizontal if cell.alignment else None,
            'vertical': cell.alignment.vertical if cell.alignment else None,
            'wrap_text': cell.alignment.wrap_text if cell.alignment else None,
            'text_rotation': cell.alignment.text_rotation if cell.alignment else None,
            'indent': cell.alignment.indent if cell.alignment else None,
        },
        'number_format': cell.number_format,
        'protection': {
            'locked': cell.protection.locked if cell.protection else None,
            'hidden': cell.protection.hidden if cell.protection else None,
        }
    }
    return style

def deep_analyze_template(filename):
    """วิเคราะห์ Template อย่างละเอียด 100%"""
    
    print(f"\n{'='*100}")
    print(f"การวิเคราะห์ Template แบบละเอียด 100%")
    print(f"ไฟล์: {filename}")
    print(f"{'='*100}\n")
    
    wb = openpyxl.load_workbook(filename, data_only=False)
    
    analysis = {
        'filename': filename,
        'sheets': []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"\n{'─'*100}")
        print(f"📋 Sheet: {sheet_name}")
        print(f"{'─'*100}\n")
        
        sheet_data = {
            'name': sheet_name,
            'dimensions': str(ws.dimensions),
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'column_widths': {},
            'row_heights': {},
            'merged_cells': [],
            'cells': {},
            'images': [],
            'sheet_properties': {
                'tabColor': ws.sheet_properties.tabColor.rgb if ws.sheet_properties.tabColor and hasattr(ws.sheet_properties.tabColor, 'rgb') else None,
            },
            'page_setup': {
                'orientation': ws.page_setup.orientation,
                'paperSize': ws.page_setup.paperSize,
                'fitToHeight': ws.page_setup.fitToHeight,
                'fitToWidth': ws.page_setup.fitToWidth,
            },
            'print_area': ws.print_area,
            'freeze_panes': str(ws.freeze_panes) if ws.freeze_panes else None,
        }
        
        # 1. Column Widths
        print("📏 ความกว้างคอลัมน์:")
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width:
                sheet_data['column_widths'][col_letter] = width
                print(f"   {col_letter}: {width}")
        
        # 2. Row Heights
        print(f"\n📏 ความสูงแถว:")
        for row_idx in range(1, ws.max_row + 1):
            height = ws.row_dimensions[row_idx].height
            if height:
                sheet_data['row_heights'][row_idx] = height
                if row_idx <= 50:  # แสดง 50 แถวแรก
                    print(f"   แถว {row_idx}: {height}")
        
        # 3. Merged Cells
        print(f"\n🔗 Merged Cells:")
        if ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                sheet_data['merged_cells'].append(str(merged_range))
                print(f"   {merged_range}")
        else:
            print("   (ไม่มี)")
        
        # 4. วิเคราะห์ทุก Cell (เฉพาะที่มีข้อมูลหรือ style)
        print(f"\n🎨 วิเคราะห์ Cell Styles (กำลังประมวลผล...)")
        cell_count = 0
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # เก็บเฉพาะ cell ที่มีข้อมูลหรือมี style
                if cell.value or cell.font.name != 'Calibri' or cell.border.left.style:
                    cell_key = f"{get_column_letter(col_idx)}{row_idx}"
                    sheet_data['cells'][cell_key] = analyze_cell_style(cell)
                    cell_count += 1
        
        print(f"   ✅ วิเคราะห์แล้ว {cell_count} cells ที่มี style")
        
        # 5. รูปภาพ (Images)
        print(f"\n🖼️ รูปภาพ:")
        if hasattr(ws, '_images') and ws._images:
            for idx, img in enumerate(ws._images):
                try:
                    img_data = {
                        'index': idx,
                        'width': img.width if hasattr(img, 'width') else None,
                        'height': img.height if hasattr(img, 'height') else None,
                    }
                    sheet_data['images'].append(img_data)
                    print(f"   รูปที่ {idx+1}: (ขนาด: {img.width}x{img.height})")
                    
                    # บันทึกรูปภาพ
                    img_filename = f"template_image_{sheet_name}_{idx+1}.png"
                    with open(img_filename, 'wb') as f:
                        f.write(img._data())
                    print(f"   💾 บันทึกเป็น: {img_filename}")
                except Exception as e:
                    print(f"   ⚠️ ไม่สามารถประมวลผลรูปที่ {idx+1}: {e}")
        else:
            print("   (ไม่มีรูปภาพ)")
        
        analysis['sheets'].append(sheet_data)
    
    # บันทึกผลการวิเคราะห์ (ข้าม JSON เนื่องจากมี RGB objects)
    print(f"\n{'='*100}")
    print(f"วิเคราะห์เสร็จสมบูรณ์!")
    print(f"จำนวน Cells ที่มี Style: {sum(len(s['cells']) for s in analysis['sheets'])}")
    print(f"{'='*100}\n")
    
    return analysis

def create_template_copy(source_file, target_file):
    """สร้าง Template ใหม่ที่เหมือนต้นฉบับ 100%"""
    
    print(f"\n{'='*100}")
    print(f"📋 กำลังสร้าง Template Copy 100%")
    print(f"{'='*100}\n")
    
    # โหลด workbook ต้นฉบับ
    wb_source = openpyxl.load_workbook(source_file)
    
    # สร้าง workbook ใหม่
    wb_target = openpyxl.Workbook()
    wb_target.remove(wb_target.active)  # ลบ sheet default
    
    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_target = wb_target.create_sheet(sheet_name)
        
        print(f"📄 Copy Sheet: {sheet_name}")
        
        # 1. Copy Column Widths
        for col_idx in range(1, ws_source.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
        
        # 2. Copy Row Heights
        for row_idx in range(1, ws_source.max_row + 1):
            ws_target.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height
        
        # 3. Copy Merged Cells
        for merged_range in ws_source.merged_cells.ranges:
            ws_target.merge_cells(str(merged_range))
        
        # 4. Copy ทุก Cell (Value + Style)
        print(f"   Copying cells...", end='', flush=True)
        for row in ws_source.iter_rows():
            for cell in row:
                target_cell = ws_target[cell.coordinate]
                
                # ข้าม merged cells (จะ copy ที่ top-left cell)
                if isinstance(cell, MergedCell):
                    continue
                
                # Copy Value
                target_cell.value = cell.value
                
                # Copy Font
                if cell.font:
                    target_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color
                    )
                
                # Copy Fill
                if cell.fill:
                    target_cell.fill = PatternFill(
                        patternType=cell.fill.patternType,
                        fgColor=cell.fill.fgColor,
                        bgColor=cell.fill.bgColor
                    )
                
                # Copy Border
                if cell.border:
                    target_cell.border = Border(
                        left=Side(style=cell.border.left.style, color=cell.border.left.color) if cell.border.left else None,
                        right=Side(style=cell.border.right.style, color=cell.border.right.color) if cell.border.right else None,
                        top=Side(style=cell.border.top.style, color=cell.border.top.color) if cell.border.top else None,
                        bottom=Side(style=cell.border.bottom.style, color=cell.border.bottom.color) if cell.border.bottom else None
                    )
                
                # Copy Alignment
                if cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                        text_rotation=cell.alignment.text_rotation,
                        indent=cell.alignment.indent
                    )
                
                # Copy Number Format
                target_cell.number_format = cell.number_format
                
                # Copy Protection
                if cell.protection:
                    target_cell.protection = Protection(
                        locked=cell.protection.locked,
                        hidden=cell.protection.hidden
                    )
        
        print(" ✅")
        
        # 5. Copy Sheet Properties
        ws_target.sheet_properties.tabColor = ws_source.sheet_properties.tabColor
        
        # 6. Copy Page Setup
        ws_target.page_setup.orientation = ws_source.page_setup.orientation
        ws_target.page_setup.paperSize = ws_source.page_setup.paperSize
        ws_target.page_setup.fitToHeight = ws_source.page_setup.fitToHeight
        ws_target.page_setup.fitToWidth = ws_source.page_setup.fitToWidth
        
        # 7. Copy Print Area
        if ws_source.print_area:
            ws_target.print_area = ws_source.print_area
        
        # 8. Copy Freeze Panes
        if ws_source.freeze_panes:
            ws_target.freeze_panes = ws_source.freeze_panes
        
        # 9. Copy Images
        if hasattr(ws_source, '_images') and ws_source._images:
            print(f"   🖼️ Copying {len(ws_source._images)} images...")
            for img in ws_source._images:
                ws_target.add_image(img)
    
    # บันทึกไฟล์
    wb_target.save(target_file)
    
    print(f"\n{'='*100}")
    print(f"✅ สร้าง Template สำเร็จ!")
    print(f"💾 บันทึกไปที่: {target_file}")
    print(f"{'='*100}\n")
    
    return target_file

if __name__ == "__main__":
    # เปลี่ยน encoding เพื่อรองรับ Thai + emoji
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    # STEP 1: วิเคราะห์ Template 100%
    print("\n" + "="*100)
    print("STEP 1: วิเคราะห์ Template อย่างละเอียด 100%")
    print("="*100)
    
    analysis = deep_analyze_template('รายงานสินค้าสำเร็จรูป.xlsx')
    
    # STEP 2: สร้าง Template Copy
    print("\n" + "="*100)
    print("STEP 2: สร้าง Template Copy ที่เหมือน 100%")
    print("="*100)
    
    target_file = create_template_copy(
        'รายงานสินค้าสำเร็จรูป.xlsx',
        'TEMPLATE-COPY-TEST.xlsx'
    )
    
    # STEP 3: ทดสอบ
    print("\n" + "="*100)
    print("STEP 3: ทดสอบ Template ที่สร้าง")
    print("="*100)
    
    print(f"\nกรุณาเปิดไฟล์ทั้ง 2 เปรียบเทียบ:")
    print(f"   1. รายงานสินค้าสำเร็จรูป.xlsx (ต้นฉบับ)")
    print(f"   2. {target_file} (Copy)")
    print(f"\nถ้าเหมือนกัน 100% = พร้อมไปขั้นตอนต่อไป!")
    print(f"\n{'='*100}\n")
