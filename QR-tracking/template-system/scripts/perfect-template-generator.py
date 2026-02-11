#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
สร้าง Template Generator ที่สมบูรณ์แบบ 100%
วิเคราะห์จาก TEMPLATE-PERFECT.xlsx และสร้างได้เหมือนกันทุกอย่าง
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from copy import copy
import json

def deep_analyze_perfect_template(source_file):
    """วิเคราะห์ Template อย่างละเอียดทุกรายละเอียด"""
    
    print(f"\n{'='*100}")
    print(f"วิเคราะห์ Template สมบูรณ์: {source_file}")
    print(f"{'='*100}\n")
    
    wb = openpyxl.load_workbook(source_file)
    
    template_data = {
        'filename': source_file,
        'sheets': {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        
        sheet_info = {
            'name': sheet_name,
            'dimensions': str(ws.dimensions),
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'column_widths': {},
            'row_heights': {},
            'merged_cells': [],
            'cells': {},
            'images': [],
            'sheet_properties': {
                'tabColor': str(ws.sheet_properties.tabColor) if ws.sheet_properties.tabColor else None,
            },
            'page_setup': {
                'orientation': ws.page_setup.orientation,
                'paperSize': ws.page_setup.paperSize,
                'fitToHeight': ws.page_setup.fitToHeight,
                'fitToWidth': ws.page_setup.fitToWidth,
            },
            'print_area': ws.print_area,
            'freeze_panes': str(ws.freeze_panes) if ws.freeze_panes else None,
        }
        
        # 1. Column Widths
        print(f"   - Analyzing column widths...")
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width:
                sheet_info['column_widths'][col_letter] = width
        
        # 2. Row Heights
        print(f"   - Analyzing row heights...")
        for row_idx in range(1, ws.max_row + 1):
            height = ws.row_dimensions[row_idx].height
            if height:
                sheet_info['row_heights'][row_idx] = height
        
        # 3. Merged Cells
        print(f"   - Analyzing merged cells...")
        if ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                sheet_info['merged_cells'].append(str(merged_range))
        
        # 4. วิเคราะห์ทุก Cell อย่างละเอียด
        print(f"   - Analyzing all cells...")
        cell_count = 0
        
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_key = f"{get_column_letter(col_idx)}{row_idx}"
                
                # เก็บข้อมูลทุก cell (รวม MergedCell เพื่อ border)
                cell_data = {
                    'coordinate': cell_key,
                    'value': cell.value,
                    'data_type': cell.data_type,
                    'is_merged': isinstance(cell, MergedCell),
                }
                
                # Font
                if cell.font:
                    cell_data['font'] = {
                        'name': cell.font.name,
                        'size': cell.font.size,
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'underline': cell.font.underline,
                        'strike': cell.font.strike,
                    }
                
                # Fill
                if cell.fill and cell.fill.patternType:
                    cell_data['fill'] = {
                        'patternType': cell.fill.patternType,
                    }
                
                # Border
                border_data = {}
                if cell.border:
                    if cell.border.left and cell.border.left.style:
                        border_data['left'] = cell.border.left.style
                    if cell.border.right and cell.border.right.style:
                        border_data['right'] = cell.border.right.style
                    if cell.border.top and cell.border.top.style:
                        border_data['top'] = cell.border.top.style
                    if cell.border.bottom and cell.border.bottom.style:
                        border_data['bottom'] = cell.border.bottom.style
                
                if border_data:
                    cell_data['border'] = border_data
                
                # Alignment
                if cell.alignment:
                    cell_data['alignment'] = {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrap_text': cell.alignment.wrap_text,
                    }
                
                # Number Format
                if cell.number_format and cell.number_format != 'General':
                    cell_data['number_format'] = cell.number_format
                
                sheet_info['cells'][cell_key] = cell_data
                cell_count += 1
        
        print(f"   - Analyzed {cell_count} cells")
        
        # 5. Images
        print(f"   - Analyzing images...")
        if hasattr(ws, '_images') and ws._images:
            for idx, img in enumerate(ws._images):
                try:
                    img_filename = f"template_perfect_img_{sheet_name}_{idx+1}.png"
                    with open(img_filename, 'wb') as f:
                        f.write(img._data())
                    
                    sheet_info['images'].append({
                        'index': idx,
                        'filename': img_filename,
                        'width': img.width if hasattr(img, 'width') else None,
                        'height': img.height if hasattr(img, 'height') else None,
                    })
                    print(f"      Saved: {img_filename}")
                except Exception as e:
                    print(f"      Error saving image {idx}: {e}")
        
        template_data['sheets'][sheet_name] = sheet_info
    
    # บันทึก analysis
    with open('TEMPLATE-PERFECT-ANALYSIS.json', 'w', encoding='utf-8') as f:
        json.dump(template_data, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n{'='*100}")
    print(f"วิเคราะห์เสร็จสมบูรณ์!")
    print(f"บันทึกไว้ที่: TEMPLATE-PERFECT-ANALYSIS.json")
    print(f"{'='*100}\n")
    
    return template_data

def create_perfect_copy(source_file, target_file):
    """สร้าง Template Copy ที่เหมือนต้นฉบับ 100%"""
    
    print(f"\n{'='*100}")
    print(f"สร้าง Perfect Copy: {source_file} -> {target_file}")
    print(f"{'='*100}\n")
    
    # โหลด source
    wb_source = openpyxl.load_workbook(source_file)
    
    # สร้าง target
    wb_target = openpyxl.Workbook()
    wb_target.remove(wb_target.active)
    
    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_target = wb_target.create_sheet(sheet_name)
        
        print(f"Copying Sheet: {sheet_name}")
        
        # 1. Column Widths
        print(f"   [1/9] Column widths...", end='', flush=True)
        for col_idx in range(1, ws_source.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
        print(" ✓")
        
        # 2. Row Heights
        print(f"   [2/9] Row heights...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            ws_target.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height
        print(" ✓")
        
        # 3. Merged Cells (ต้อง merge ก่อน copy cells)
        print(f"   [3/9] Merged cells...", end='', flush=True)
        for merged_range in ws_source.merged_cells.ranges:
            ws_target.merge_cells(str(merged_range))
        print(" ✓")
        
        # 4. Cell Values (เฉพาะ non-merged cells) - copy ทุก value รวม empty string
        print(f"   [4/9] Cell values...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                # Copy value ทุก cell ที่ไม่ใช่ MergedCell
                if not isinstance(source_cell, MergedCell):
                    target_cell.value = source_cell.value
        print(" ✓")
        
        # 5. Font (ทุก cell)
        print(f"   [5/9] Fonts...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
        print(" ✓")
        
        # 6. Fill (ทุก cell)
        print(f"   [6/9] Fills...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
        print(" ✓")
        
        # 7. Border (ทุก cell - สำคัญมาก!)
        print(f"   [7/9] Borders...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
        print(" ✓")
        
        # 8. Alignment & Number Format (ทุก cell)
        print(f"   [8/9] Alignment & Format...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
        print(" ✓")
        
        # 9. Images
        print(f"   [9/9] Images...", end='', flush=True)
        if hasattr(ws_source, '_images') and ws_source._images:
            for img in ws_source._images:
                ws_target.add_image(img)
            print(f" ✓ ({len(ws_source._images)} images)")
        else:
            print(" ✓ (no images)")
        
        # 10. แก้ไข checkbox ในแถว 27 (เปลี่ยน o เป็น □)
        print(f"   [10/10] Fix checkbox row 27...", end='', flush=True)
        cell_a27 = ws_target['A27']
        if cell_a27.value and isinstance(cell_a27.value, str):
            # เปลี่ยน "o  ปนเปื้อน" เป็น "□  ปนเปื้อน" และ "o  ไม่ปนเปื้อน" เป็น "□  ไม่ปนเปื้อน"
            text = cell_a27.value
            text = text.replace('o  ปนเปื้อน', '□  ปนเปื้อน')
            text = text.replace('o  ไม่ปนเปื้อน', '□  ไม่ปนเปื้อน')
            cell_a27.value = text
            print(" ✓")
        else:
            print(" (not found)")
        
        # Sheet Properties
        ws_target.sheet_properties.tabColor = ws_source.sheet_properties.tabColor
        ws_target.page_setup.orientation = ws_source.page_setup.orientation
        ws_target.page_setup.paperSize = ws_source.page_setup.paperSize
        ws_target.page_setup.fitToHeight = ws_source.page_setup.fitToHeight
        ws_target.page_setup.fitToWidth = ws_source.page_setup.fitToWidth
        
        if ws_source.print_area:
            ws_target.print_area = ws_source.print_area
        
        if ws_source.freeze_panes:
            ws_target.freeze_panes = ws_source.freeze_panes
    
    # Save
    wb_target.save(target_file)
    
    print(f"\n{'='*100}")
    print(f"สร้างเสร็จสมบูรณ์: {target_file}")
    print(f"{'='*100}\n")
    
    return target_file

def compare_templates(file1, file2):
    """เปรียบเทียบ 2 ไฟล์ว่าเหมือนกันหรือไม่"""
    
    print(f"\n{'='*100}")
    print(f"เปรียบเทียบ: {file1} vs {file2}")
    print(f"{'='*100}\n")
    
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    
    differences = []
    
    for sheet_name in wb1.sheetnames:
        if sheet_name not in wb2.sheetnames:
            differences.append(f"Sheet {sheet_name} ไม่มีในไฟล์ที่ 2")
            continue
        
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        
        # ตรวจสอบ dimensions
        if ws1.max_row != ws2.max_row or ws1.max_column != ws2.max_column:
            diff = f"   ✗ Dimensions: {ws1.max_row}x{ws1.max_column} vs {ws2.max_row}x{ws2.max_column}"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Dimensions: {ws1.max_row}x{ws1.max_column}")
        
        # ตรวจสอบ column widths
        col_diff = 0
        for col_idx in range(1, ws1.max_column + 1):
            col_letter = get_column_letter(col_idx)
            w1 = ws1.column_dimensions[col_letter].width
            w2 = ws2.column_dimensions[col_letter].width
            if abs(w1 - w2) > 0.01:
                col_diff += 1
        
        if col_diff > 0:
            diff = f"   ✗ Column widths differ: {col_diff} columns"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Column widths: identical")
        
        # ตรวจสอบ row heights
        row_diff = 0
        for row_idx in range(1, ws1.max_row + 1):
            h1 = ws1.row_dimensions[row_idx].height
            h2 = ws2.row_dimensions[row_idx].height
            if h1 and h2 and abs(h1 - h2) > 0.01:
                row_diff += 1
        
        if row_diff > 0:
            diff = f"   ✗ Row heights differ: {row_diff} rows"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Row heights: identical")
        
        # ตรวจสอบ merged cells
        merged1 = set(str(r) for r in ws1.merged_cells.ranges)
        merged2 = set(str(r) for r in ws2.merged_cells.ranges)
        
        if merged1 != merged2:
            diff = f"   ✗ Merged cells differ: {len(merged1)} vs {len(merged2)}"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Merged cells: {len(merged1)} ranges")
        
        # ตรวจสอบ cell values
        value_diff = 0
        for row_idx in range(1, min(ws1.max_row, ws2.max_row) + 1):
            for col_idx in range(1, min(ws1.max_column, ws2.max_column) + 1):
                c1 = ws1.cell(row=row_idx, column=col_idx)
                c2 = ws2.cell(row=row_idx, column=col_idx)
                
                if not isinstance(c1, MergedCell) and not isinstance(c2, MergedCell):
                    # ถือว่า empty string และ None เป็นสิ่งเดียวกัน
                    v1 = c1.value if c1.value != '' else None
                    v2 = c2.value if c2.value != '' else None
                    
                    if v1 != v2:
                        value_diff += 1
        
        if value_diff > 0:
            diff = f"   ✗ Cell values differ: {value_diff} cells"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Cell values: identical")
        
        # ตรวจสอบ borders
        border_diff = 0
        for row_idx in range(1, min(ws1.max_row, ws2.max_row) + 1):
            for col_idx in range(1, min(ws1.max_column, ws2.max_column) + 1):
                c1 = ws1.cell(row=row_idx, column=col_idx)
                c2 = ws2.cell(row=row_idx, column=col_idx)
                
                b1 = (c1.border.left.style if c1.border and c1.border.left else None,
                      c1.border.right.style if c1.border and c1.border.right else None,
                      c1.border.top.style if c1.border and c1.border.top else None,
                      c1.border.bottom.style if c1.border and c1.border.bottom else None)
                
                b2 = (c2.border.left.style if c2.border and c2.border.left else None,
                      c2.border.right.style if c2.border and c2.border.right else None,
                      c2.border.top.style if c2.border and c2.border.top else None,
                      c2.border.bottom.style if c2.border and c2.border.bottom else None)
                
                if b1 != b2:
                    border_diff += 1
        
        if border_diff > 0:
            diff = f"   ✗ Borders differ: {border_diff} cells"
            differences.append(diff)
            print(diff)
        else:
            print(f"   ✓ Borders: identical")
        
        print()
    
    print(f"{'='*100}")
    if differences:
        print(f"พบความแตกต่าง {len(differences)} จุด:")
        for d in differences:
            print(f"  {d}")
    else:
        print(f"✓✓✓ ไฟล์ทั้ง 2 เหมือนกัน 100%! ✓✓✓")
    print(f"{'='*100}\n")
    
    return len(differences) == 0

if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    source = 'TEMPLATE-PERFECT.xlsx'
    target = 'TEMPLATE-GENERATED.xlsx'
    
    # Step 1: วิเคราะห์
    print("\n" + "="*100)
    print("STEP 1: วิเคราะห์ Template สมบูรณ์")
    print("="*100)
    template_data = deep_analyze_perfect_template(source)
    
    # Step 2: สร้าง Copy
    print("\n" + "="*100)
    print("STEP 2: สร้าง Perfect Copy")
    print("="*100)
    create_perfect_copy(source, target)
    
    # Step 3: เปรียบเทียบ
    print("\n" + "="*100)
    print("STEP 3: เปรียบเทียบความเหมือน")
    print("="*100)
    is_identical = compare_templates(source, target)
    
    # สรุป
    print("\n" + "="*100)
    print("สรุปผลการทำงาน")
    print("="*100)
    print(f"Source: {source}")
    print(f"Generated: {target}")
    print(f"Analysis: TEMPLATE-PERFECT-ANALYSIS.json")
    
    if is_identical:
        print(f"\n✓✓✓ สำเร็จ! ไฟล์ที่สร้างเหมือนต้นฉบับ 100% ✓✓✓")
    else:
        print(f"\n✗ พบความแตกต่าง กรุณาตรวจสอบ")
    
    print("="*100 + "\n")
