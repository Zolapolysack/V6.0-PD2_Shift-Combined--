#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
หา cell ที่แตกต่างกัน
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

def find_different_cells(file1, file2):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)
    
    for sheet_name in wb1.sheetnames:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        
        print(f"\nSheet: {sheet_name}\n")
        
        for row_idx in range(1, ws1.max_row + 1):
            for col_idx in range(1, ws1.max_column + 1):
                c1 = ws1.cell(row=row_idx, column=col_idx)
                c2 = ws2.cell(row=row_idx, column=col_idx)
                
                if not isinstance(c1, MergedCell) and not isinstance(c2, MergedCell):
                    if c1.value != c2.value:
                        print(f"Cell {c1.coordinate}:")
                        print(f"   File 1: {repr(c1.value)}")
                        print(f"   File 2: {repr(c2.value)}")
                        print()

if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    find_different_cells('TEMPLATE-PERFECT.xlsx', 'TEMPLATE-GENERATED.xlsx')
