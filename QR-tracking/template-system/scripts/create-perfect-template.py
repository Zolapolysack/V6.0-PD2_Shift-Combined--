#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
สร้าง Template Copy 100% พร้อมแก้ไข:
1. Border ให้ครบทุก cell
2. เพิ่ม Checkbox (☐) หน้าข้อความ
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

def create_perfect_template(source_file, target_file):
    """สร้าง Template ที่สมบูรณ์ 100%"""
    
    print(f"\n{'='*100}")
    print(f"กำลังสร้าง Template Copy 100% (พร้อมแก้ไข Border และ Checkbox)")
    print(f"{'='*100}\n")
    
    # โหลด workbook ต้นฉบับ
    wb_source = openpyxl.load_workbook(source_file)
    
    # สร้าง workbook ใหม่
    wb_target = openpyxl.Workbook()
    wb_target.remove(wb_target.active)
    
    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_target = wb_target.create_sheet(sheet_name)
        
        print(f"Copy Sheet: {sheet_name}")
        
        # 1. Copy Column Widths
        print("   - Column widths...", end='', flush=True)
        for col_idx in range(1, ws_source.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
        print(" OK")
        
        # 2. Copy Row Heights
        print("   - Row heights...", end='', flush=True)
        for row_idx in range(1, ws_source.max_row + 1):
            ws_target.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height
        print(" OK")
        
        # 3. Copy Merged Cells (ก่อน copy cells)
        print("   - Merged cells...", end='', flush=True)
        for merged_range in ws_source.merged_cells.ranges:
            ws_target.merge_cells(str(merged_range))
        print(" OK")
        
        # 4. Copy ทุก Cell (Value + Style) - รวมถึง MergedCells
        print("   - Cell contents and styles...", end='', flush=True)
        
        # Copy ทุก cell รวมถึง merged cells เพื่อให้ border ครบ
        for row_idx in range(1, ws_source.max_row + 1):
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                target_cell = ws_target.cell(row=row_idx, column=col_idx)
                
                # Copy Value (ยกเว้น MergedCell ที่ไม่ใช่ top-left)
                if not isinstance(source_cell, MergedCell):
                    if source_cell.value:
                        target_cell.value = source_cell.value
                
                # Copy Styles ทุก cell (รวม MergedCell เพื่อให้ border ครบ)
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                
                # Border - สำคัญที่สุด! copy ทุก cell
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
        
        print(" OK")
        
        # 5. เพิ่ม Checkbox ให้ถูกต้อง
        print("   - Fixing checkboxes...", end='', flush=True)
        checkbox_fixed = 0
        
        # เปลี่ยน ¨ เป็น £ ให้เหมือนต้นฉบับ
        for row in ws_target.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    text = str(cell.value)
                    
                    # แก้ไข ¨ พิมพ์ และ ¨ ไม่พิมพ์ เป็น £
                    if '¨   พิมพ์' in text:
                        cell.value = text.replace('¨   พิมพ์', '£   พิมพ์')
                        checkbox_fixed += 1
                    elif '¨   ไม่พิมพ์' in text:
                        cell.value = text.replace('¨   ไม่พิมพ์', '£   ไม่พิมพ์')
                        checkbox_fixed += 1
        
        print(f" OK (แก้ไข {checkbox_fixed} checkboxes)")
        
        # 6. Copy Images
        print("   - Images...", end='', flush=True)
        if hasattr(ws_source, '_images') and ws_source._images:
            for img in ws_source._images:
                ws_target.add_image(img)
            print(f" OK ({len(ws_source._images)} images)")
        else:
            print(" OK (no images)")
        
        # 7. Copy Sheet Properties
        ws_target.sheet_properties.tabColor = ws_source.sheet_properties.tabColor
        ws_target.page_setup.orientation = ws_source.page_setup.orientation
        ws_target.page_setup.paperSize = ws_source.page_setup.paperSize
        
        if ws_source.print_area:
            ws_target.print_area = ws_source.print_area
        
        if ws_source.freeze_panes:
            ws_target.freeze_panes = ws_source.freeze_panes
    
    # บันทึกไฟล์
    wb_target.save(target_file)
    
    print(f"\n{'='*100}")
    print(f"สร้าง Template สำเร็จ!")
    print(f"บันทึกไปที่: {target_file}")
    print(f"{'='*100}\n")
    
    return target_file

def verify_template(filename):
    """ตรวจสอบ Template ที่สร้าง"""
    
    print(f"\n{'='*100}")
    print(f"ตรวจสอบ Template: {filename}")
    print(f"{'='*100}\n")
    
    wb = openpyxl.load_workbook(filename)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        
        # นับ cells ที่มี border
        border_count = 0
        checkbox_count = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                
                # นับ border
                if cell.border and (cell.border.left.style or cell.border.right.style or 
                                   cell.border.top.style or cell.border.bottom.style):
                    border_count += 1
                
                # นับ checkbox
                if cell.value and isinstance(cell.value, str):
                    if '☐' in str(cell.value) or '□' in str(cell.value):
                        checkbox_count += 1
        
        print(f"   - Cells with borders: {border_count}")
        print(f"   - Checkboxes found: {checkbox_count}")
        print()
    
    print(f"{'='*100}\n")

if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    # สร้าง Template
    target_file = create_perfect_template(
        'รายงานสินค้าสำเร็จรูป.xlsx',
        'TEMPLATE-PERFECT.xlsx'
    )
    
    # ตรวจสอบ
    verify_template(target_file)
    
    print("=" * 100)
    print("เสร็จสิ้น! กรุณาเปิดไฟล์ TEMPLATE-PERFECT.xlsx เพื่อตรวจสอบ")
    print("ตรวจสอบ 2 จุด:")
    print("   1. เส้นขอบครบถ้วนหรือไม่")
    print("   2. มี checkbox (☐) หน้าคำว่า ผ่าน, ไม่ผ่าน, ปนเปื้อน, ฯลฯ หรือไม่")
    print("=" * 100)
