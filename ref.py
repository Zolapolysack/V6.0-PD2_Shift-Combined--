!pip install pandas openpyxl xlsxwriter > /dev/null
import re
import io
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from zoneinfo import ZoneInfo
from google.colab import files
import warnings

# --------------------------------
# กำหนดค่าคงที่
# --------------------------------
PRIMARY_MACHINE_ORDER = [
    'CL1','CL2','CL3','CL4','CL5','CL6','CL7','CL8','CL9','CL12','CL13','CL14','CL15',
    'CL16','CL17','CL19','CL25','CL26','CL27','CL28','CL29','CL30','CL31','CL32','CL33',
    'CL44','CL45','CL46','CL47','CL48','CL49','CL50','CL51','CL52','CL53','CL54','CL55',
    'CL56','CL57','CL58','CL59','CL60','CL61','CL62','CL63','CL64','CL65','CL66','CL67',
    'CL68','CL69','CL70','CL71','CL72','CL73','CL74','CL75','CL76','CL77','CL78','CL79',
    'CL80','CL81','CL82','CL83','CL84','CL85','CL86','CL87','CL88','CL89','CL90','CL91',
    'CL92','CL93','CL94','CL95','CL34','CL35','CL36'
 ]
# รายการเลขเครื่องสำหรับตัดม้วนครั้งที่ 2 และ 3
CUT2_LIST = [m + " ตัดม้วนครั้งที่ 2" for m in PRIMARY_MACHINE_ORDER]
CUT3_LIST = [m + " ตัดม้วนครั้งที่ 3" for m in PRIMARY_MACHINE_ORDER]
EXTRA_MACHINE_ORDER_FULL = CUT2_LIST + CUT3_LIST

# กำหนดค่าตำแหน่งการแสดงผลใน Excel
EXTRA_BLOCK_HEADER_EXCEL_ROW = 109
EXTRA_BLOCK_DATA_START_EXCEL_ROW = 110
MAX_ROW_FOR_BORDER_PRIMARY = 83

# โครงสร้างคอลัมน์ที่สร้างเมื่ออัพโหลดไฟล์เข้ามา (กะ A/B)
EXPECTED_COLUMN_COUNT_AB = 21
RAW_COLUMNS_AB = [
    'เวลาบันทึก','ลำดับชุดข้อมูลที่','วันที่','เครื่องทอ NO','พนักงานทอ กะ A','พนักงานทอ กะ B',
    'ขนาดหน้าผ้า_รหัส','Lot.No','เลขที่ใบสั่งผลิต','ยอดทอ กะ A','ยอดทอ กะ B',
    'ความเร็วรอบ กะ A','ความเร็วรอบ กะ B','ประสิทธิภาพ กะ A','ประสิทธิภาพ กะ B',
    'มิเตอร์เริ่มงาน กะ A','มิเตอร์เลิกงาน กะ A','มิเตอร์เริ่มงาน กะ B','มิเตอร์เลิกงาน กะ B',
    'ความยาวตัดม้วน กะ A','ความยาวตัดม้วน กะ B',
]

# โครงสร้างคอลัมน์สำหรับไฟล์เสริม (ไฟล์ C)
EXPECTED_COLUMN_COUNT_C = 8
RAW_COLUMNS_C = [
    'เวลาบันทึก','ลำดับชุดข้อมูลที่','วันที่','เครื่องทอ NO','ขนาดหน้าผ้า (รหัส)',
    'ขนาด Lot.No','ความยาวตัดม้วน กะ A','ความยาวตัดม้วน กะ B'
 ]

# คอลัมน์ที่เป็นตัวเลข (ใช้กำหนดความกว้าง/ฟอร์แมต)
NUMERIC_DATA_COLUMNS = [
    "ยอดทอ กะ A","ยอดทอ กะ B","ความเร็วรอบ กะ A","ความเร็วรอบ กะ B",
    "ประสิทธิภาพ (กะ A)%","ประสิทธิภาพ (กะ B)%","มิเตอร์เริ่มงาน กะ A","มิเตอร์เลิกงาน กะ A",
    "มิเตอร์เริ่มงาน กะ B","มิเตอร์เลิกงาน กะ B","ความยาวตัดม้วน กะ A (เดิม)","ความยาวตัดม้วน กะ B (เดิม)",
    "Y","Z",
]

# ---------------------------------------------------------------------------
# Mapping codes (Case-Insensitive)
# ---------------------------------------------------------------------------
# หมายเหตุ: ปรับให้ lookup ไม่สนตัวพิมพ์ (ทั้งพิมพ์เล็ก/ใหญ่/มีช่องว่าง)
CODE_MAPPING_RAW = {
    '1625800':'01','1825800':'02','1825800SM':'03','1925800':'04','1925800SM':'05',
    '1921720':'06','2025700':'07','2025800':'08','2025800SM':'09','2021750':'10',
    '2125650':'11','2121670SM':'12','2325650':'13','1425800':'14','2625650':'15',
    '262575051':'16','1725800':'17','2625800':'18','2025850':'19','2221940':'20',
    'SCL0223211125':'21','SCL052325650':'22','2321750':'23','2421112512551':'24',
    '20251000':'25','2625850SM':'26','2325800SM':'27','2825850':'28','2125800SMN':'29',
    '19251000sm':'30','1625800SM':'31','2021720UV':'32','23211125130':'33',
    '2321112512551':'34','2321112514051':'35','2025650':'36','2125700':'37',
    '2125800':'38','2021850UV':'39','1R042325650':'40','1621720SMFX':'41',
    'FCL022325650':'42','23211125120smfx':'43',
}
# เพิ่ม synonym/รูปแบบสะกดอื่น
EXTRA_CODE_SYNONYMS = {
    '23211125120SMFX': '43',
    '1925100SM': '30',
    '19251000SM': '30',
    '2321112512551SMFX':'34'
}
for _k,_v in EXTRA_CODE_SYNONYMS.items():
    CODE_MAPPING_RAW[_k] = _v


def normalize_code(val: str) -> str:
    """แปลงรหัสให้พร้อม lookup: strip -> remove inner spaces -> upper"""
    if val is None:
        return ''
    return str(val).strip().replace(' ', '').upper()


# สร้าง mapping ที่ normalize แล้ว (คีย์ทั้งหมดเป็น UPPERCASE ไม่มี space)
CODE_MAPPING = {normalize_code(k): v for k, v in CODE_MAPPING_RAW.items()}


def map_code(raw_value: str):
    """คืนค่ารหัส 2 หลัก ตาม mapping (ไม่สนตัวพิมพ์/space); ถ้าไม่พบคืน None"""
    key = normalize_code(raw_value)
    if not key:
        return None
    v = CODE_MAPPING.get(key)
    if v is None and key.endswith('SM') and key + '0' in CODE_MAPPING:
        v = CODE_MAPPING.get(key + '0')
    return v

# ---------------------------------------------------------------------------
# ฟังก์ชันกำหนดเลขที่ใบสั่งผลิต (ใช้ mapping  แบบไม่สนตัวพิมพ์ (ทั้งพิมพ์เล็ก/ใหญ่/มีช่องว่าง))
# ---------------------------------------------------------------------------

def assign_production_code(df: pd.DataFrame) -> pd.DataFrame:
    """กำหนดค่าเลขที่ใบสั่งผลิต = YYMM + XX (XX จาก mapping case-insensitive)"""
    target_col_name = 'เลขที่ใบสั่งผลิต'
    possible_source_cols = ['ขนาดหน้าผ้า (รหัส)', 'ขนาดหน้าผ้า_รหัส']
    source_col_name = None
    for c in possible_source_cols:
        if c in df.columns:
            source_col_name = c
            break
    if target_col_name not in df.columns:
        df[target_col_name] = ''
    df[target_col_name] = df[target_col_name].astype(object)
    if source_col_name is None:
        return df
    try:
        now = datetime.now(ZoneInfo('Asia/Bangkok'))
    except Exception:
        now = datetime.now()
    buddhist_year = now.year + 543
    prefix = f"{buddhist_year % 100:02d}{now.month:02d}"
    for idx, raw_val in df[source_col_name].items():
        code2 = map_code(raw_val)
        if code2:
            df.at[idx, target_col_name] = str(prefix + code2)
    return df


# -------------------------------------------------------------
# ฟังก์ชันเครื่องมือ: อ่านไฟล์ จัดหัวตาราง
# -------------------------------------------------------------

def find_header_row(df, keyword='เครื่องทอ NO'):
    for i, row in df.iterrows():
        s = row.astype(str)
        try:
            if s.str.contains(keyword, na=False).any():
                return i
        except Exception:
            pass
        found = False
        for val in s:
            if not isinstance(val, str):
                continue
            v = val.strip()
            if 'เครื่อง' in v and 'NO' in v.upper():
                found = True
                break
        if found:
            return i
    return -1


def _parse_date_series(series: pd.Series) -> pd.Series:
    """Detect a dominant single date format and parse consistently.
    Steps:
    1. Clean text & split off time part.
    2. Convert Buddhist years (>2400) to Gregorian (year-543).
    3. Try candidate patterns; pick the one with max successes (>=1).
    4. Parse ONLY with that format (no infer_datetime_format) -> consistency, no warnings.
    If none match, final fallback per-element parse with dayfirst, but without deprecated arg.
    """
    if series is None or series.empty:
        return pd.to_datetime(series, errors='coerce')
    raw = series.astype(str).str.strip()
    # Drop obvious blanks
    raw = raw.replace({'': np.nan, 'None': np.nan, 'NaT': np.nan})
    # Remove time part if present (split at first space if pattern has colon)
    cleaned = raw.str.split().str[0]


    # Normalize separators and zero-pad day/month if needed later
    # Convert Buddhist year if numeric token length 4 and > 2400
    def adjust_buddhist(token):
        if re.fullmatch(r'\d{4}', token):
            yr = int(token)
            if yr > 2400:
                return str(yr - 543)
        return token


    # Candidate patterns with a regex builder and corresponding strptime format
    candidates = [
        (re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{4})$'), '%d/%m/%Y'),
        (re.compile(r'^(\d{1,2})-(\d{1,2})-(\d{4})$'), '%d-%m-%Y'),
        (re.compile(r'^(\d{4})-(\d{1,2})-(\d{1,2})$'), '%Y-%m-%d'),
        (re.compile(r'^(\d{4})/(\d{1,2})/(\d{1,2})$'), '%Y/%m/%d'),
        (re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{2})$'), '%d/%m/%y'),
    ]


    # Pre-normalize each value for buddhist year and keep mapping
    norm_values = []
    for v in cleaned.fillna('').tolist():
        if not v:
            norm_values.append('')
            continue
        parts = re.split(r'[-/]', v)
        parts = [adjust_buddhist(p) for p in parts]
        if len(parts) == 3:
            sep = '-' if '-' in v else ('/' if '/' in v else '/')
            v2 = sep.join(parts)
        else:
            v2 = v
        norm_values.append(v2)
    norm_ser = pd.Series(norm_values, index=series.index)


    best_fmt = None
    best_mask = None
    for regex, fmt in candidates:
        mask = norm_ser.str.match(regex)
        if mask.any():
            # success count
            if best_fmt is None or mask.sum() > best_mask.sum():
                best_fmt = fmt
                best_mask = mask
    if best_fmt:
        parsed = pd.to_datetime(norm_ser.where(best_mask), format=best_fmt, errors='coerce')
        # Non-matching -> NaT (consistent)
        return parsed
    # Fallback: dayfirst=True without deprecated infer flag
    try:
        return pd.to_datetime(norm_ser, dayfirst=True, errors='coerce')
    except Exception:
        return pd.to_datetime(norm_ser, errors='coerce')


def read_generic_file(file_content, filename, expected_cols, col_names):
    df = None
    ext = '.' + filename.split('.')[-1].lower()
    try:
        if ext == '.xlsx':
            df = pd.read_excel(io.BytesIO(file_content), header=None)
        elif ext == '.csv':
            for enc in ['utf-8','tis-620','cp874','latin1']:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), header=None,
                                     on_bad_lines='skip', encoding=enc)
                    break
                except Exception:
                    continue
    except Exception as e:
        print(f"[อ่านไฟล์ไม่สำเร็จ] {filename} : {e}")
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    header_idx = find_header_row(df)
    if header_idx != -1:
        df.columns = df.iloc[header_idx]
        df = df.iloc[header_idx+1:].reset_index(drop=True)
    else:
        if df.shape[1] > len(col_names):
            df = df.iloc[:, :len(col_names)]
        df.columns = col_names[:df.shape[1]]


    def _normalize_col(s: object) -> str:
        if s is None:
            return ''
        txt = str(s).strip()
        txt = re.sub(r"\s+", " ", txt)
        up = txt.upper()
        if 'เครื่อง' in txt and 'NO' in up:
            return 'เครื่องทอ NO'
        if 'ขนาด' in txt and 'หน้าผ้า' in txt:
            if 'รหัส' in txt or 'CODE' in up:
                return 'ขนาดหน้าผ้า (รหัส)'
            else:
                return 'ขนาดหน้าผ้า_รหัส'
        return txt
    df.rename(columns=_normalize_col, inplace=True)


    if df.shape[1] < expected_cols:
        df.columns = col_names[:df.shape[1]]
    else:
        df = df.iloc[:, :expected_cols]
        df.columns = col_names


    if 'วันที่' in df.columns:
        df['วันที่'] = _parse_date_series(df['วันที่'])
    if 'เครื่องทอ NO' in df.columns:
        df['เครื่องทอ NO'] = df['เครื่องทอ NO'].astype(str).str.strip()


    total = len(df)
    if total > 0:
        missing_date = df['วันที่'].isna().sum() if 'วันที่' in df.columns else total
        missing_machine = df['เครื่องทอ NO'].isna().sum() if 'เครื่องทอ NO' in df.columns else total
        pct_missing_machine = missing_machine / total
        if pct_missing_machine > 0.5:
            raise ValueError(f"ไฟล์ {filename}: มากกว่า 50% ของแถวไม่มีค่า 'เครื่องทอ NO' ({missing_machine}/{total})")
        if missing_date / total > 0.5:
            warnings.warn(f"ไฟล์ {filename}: มากกว่า 50% ของแถวไม่สามารถแปลง 'วันที่' ได้ ({missing_date}/{total}). ผลการเรียง/เลือกแถวอาจผิดพลาด", UserWarning)


    if {'วันที่', 'เครื่องทอ NO'}.issubset(df.columns):
        df.dropna(subset=['วันที่', 'เครื่องทอ NO'], inplace=True)
    return df


def extract_machine_number(s):
    if not isinstance(s, str):
        return 10**9
    m = re.search(r'CL(\d+)', s.upper())
    return int(m.group(1)) if m else 10**9


def detect_cut_iteration(s):
    if not isinstance(s, str):
        return 0
    if "ตัดม้วนที่ครั้ง 2" in s or "ตัดม้วนครั้งที่ 2" in s:
        return 2
    if "ตัดม้วนที่ครั้ง 3" in s or "ตัดม้วนครั้งที่ 3" in s:
        return 3
    return 0


def generate_and_assign_lot_no(df, run_date=None, max_sequence=999):
    full_machine_order = PRIMARY_MACHINE_ORDER + CUT2_LIST + CUT3_LIST
    machine_to_order_map = {m: i for i, m in enumerate(full_machine_order)}
    a_vals = pd.to_numeric(
        df.get('ความยาวตัดม้วน กะ A (เดิม)', pd.Series(dtype=float)),
        errors='coerce'
    ).fillna(0)
    b_vals = pd.to_numeric(
        df.get('ความยาวตัดม้วน กะ B (เดิม)', pd.Series(dtype=float)),
        errors='coerce'
    ).fillna(0)
    needs = (a_vals > 0) | (b_vals > 0)
    df['Lot.No'] = df.get('Lot.No', '')
    if not needs.any():
        df.loc[~needs, 'Lot.No'] = ''
        return df
    now = run_date if run_date is not None else datetime.now(ZoneInfo('Asia/Bangkok'))
    buddhist_year = now.year + 543
    date_prefix = now.strftime(f"{str(buddhist_year)[-2:]}%m%d")
    candidates = df[needs].copy()
    candidates['_order_idx'] = candidates['เครื่อง'].map(
        lambda v: machine_to_order_map.get(v, float('inf'))
)
    candidates['_num'] = candidates['เครื่อง'].map(lambda v: extract_machine_number(v))
    candidates.sort_values(['_order_idx', '_num'], inplace=True)
    seq = 1
    for idx in candidates.index:
        if seq > max_sequence:
            raise ValueError(
                f"sequence number exceeded max_sequence={max_sequence}"
)
        df.at[idx, 'Lot.No'] = f"{date_prefix}{seq:02d}" if max_sequence < 1000 else f"{date_prefix}{seq:02d}"
        seq += 1
    return df


def _ensure_cols(df_in):
    df = df_in.copy()
    if 'วันที่' not in df.columns:
        df['วันที่'] = pd.NaT
    if 'เครื่องทอ NO' not in df.columns:
        for c in list(df.columns):
            try:
                if isinstance(c, str) and 'เครื่อง' in c and 'NO' in c.upper():
                    df.rename(columns={c: 'เครื่องทอ NO'}, inplace=True)
                    print(f"mapped column '{c}' -> 'เครื่องทอ NO'")
                    break
            except Exception:
                continue
    if 'เครื่องทอ NO' not in df.columns:
        df['เครื่องทอ NO'] = ''
    return df


def build_primary_block(df_a_raw, df_b_raw, df_c_raw):
    def _ensure_cols(df_in):
        df = df_in.copy()
        if 'วันที่' not in df.columns:
            df['วันที่'] = pd.NaT
        if 'เครื่องทอ NO' not in df.columns:
            for c in list(df.columns):
                try:
                    if isinstance(c, str) and 'เครื่อง' in c and 'NO' in c.upper():
                        df.rename(columns={c: 'เครื่องทอ NO'}, inplace=True)
                        print(f"mapped column '{c}' -> 'เครื่องทอ NO'")
                        break
                except Exception:
                    continue
        if 'เครื่องทอ NO' not in df.columns:
            df['เครื่องทอ NO'] = ''
        return df
    df_a = (_ensure_cols(df_a_raw)
            .sort_values(by='วันที่', ascending=False)
            .drop_duplicates('เครื่องทอ NO', keep='first'))
    df_b = (_ensure_cols(df_b_raw)
            .sort_values(by='วันที่', ascending=False)
            .drop_duplicates('เครื่องทอ NO', keep='first'))
    cols_a = [
        'เครื่องทอ NO','พนักงานทอ กะ A','ขนาดหน้าผ้า_รหัส','Lot.No','เลขที่ใบสั่งผลิต',
        'ยอดทอ กะ A','ความเร็วรอบ กะ A','ประสิทธิภาพ กะ A','มิเตอร์เริ่มงาน กะ A',
        'มิเตอร์เลิกงาน กะ A','ความยาวตัดม้วน กะ A'
    ]
    cols_b = [
        'เครื่องทอ NO','พนักงานทอ กะ B','ขนาดหน้าผ้า_รหัส','ยอดทอ กะ B','ความเร็วรอบ กะ B',
        'ประสิทธิภาพ กะ B','มิเตอร์เริ่มงาน กะ B','มิเตอร์เลิกงาน กะ B','ความยาวตัดม้วน กะ B'
    ]
    merged = pd.merge(df_a[cols_a], df_b[cols_b], on='เครื่องทอ NO', how='outer')
    if 'ขนาดหน้าผ้า_รหัส_x' in merged.columns:
        merged['ขนาดหน้าผ้า_รหัส'] = merged['ขนาดหน้าผ้า_รหัส_x'].combine_first(
            merged['ขนาดหน้าผ้า_รหัส_y']
        )
        merged.drop(columns=['ขนาดหน้าผ้า_รหัส_x','ขนาดหน้าผ้า_รหัส_y'], inplace=True)
    if not df_c_raw.empty and 'เครื่องทอ NO' in df_c_raw.columns:
        df_c = (_ensure_cols(df_c_raw)
                .sort_values(by='วันที่', ascending=False)
                .drop_duplicates('เครื่องทอ NO', keep='first'))
        df_c = df_c[
            ['เครื่องทอ NO','ขนาด Lot.No','ความยาวตัดม้วน กะ A','ความยาวตัดม้วน กะ B']
        ].rename(columns={
            'ขนาด Lot.No':'Lot.No_C',
            'ความยาวตัดม้วน กะ A':'Y',
            'ความยาวตัดม้วน กะ B':'Z',
        })
        merged = pd.merge(merged, df_c, on='เครื่องทอ NO', how='left')
        merged['Y'] = merged.get('Y')
        merged['Z'] = merged.get('Z')
    else:
        merged['Y'] = np.nan
        merged['Z'] = np.nan
    scaffold = pd.DataFrame(PRIMARY_MACHINE_ORDER, columns=['เครื่องทอ NO'])
    final = pd.merge(scaffold, merged, on='เครื่องทอ NO', how='left')
    export_df = pd.DataFrame({
        "เครื่อง": final['เครื่องทอ NO'],
        "พนักงานทอ กะ A": final.get('พนักงานทอ กะ A'),
        "พนักงานทอ กะ B": final.get('พนักงานทอ กะ B'),
        "ขนาดหน้าผ้า (รหัส)": final.get('ขนาดหน้าผ้า_รหัส'),
        "Lot.No": final.get('Lot.No'),
        "เลขที่ใบสั่งผลิต": final.get('เลขที่ใบสั่งผลิต'),
        "ยอดทอ กะ A": np.nan,
        "ยอดทอ กะ B": np.nan,
        "ความเร็วรอบ กะ A": final.get('ความเร็วรอบ กะ A'),
        "ความเร็วรอบ กะ B": final.get('ความเร็วรอบ กะ B'),
        "ประสิทธิภาพ (กะ A)%": final.get('ประสิทธิภาพ กะ A'),
        "ประสิทธิภาพ (กะ B)%": final.get('ประสิทธิภาพ กะ B'),
        "มิเตอร์เริ่มงาน กะ A": final.get('มิเตอร์เริ่มงาน กะ A'),
        "มิเตอร์เลิกงาน กะ A": final.get('มิเตอร์เลิกงาน กะ A'),
        "มิเตอร์เริ่มงาน กะ B": final.get('มิเตอร์เริ่มงาน กะ B'),
        "มิเตอร์เลิกงาน กะ B": final.get('มิเตอร์เลิกงาน กะ B'),
        "ความยาวตัดม้วน กะ A (เดิม)": final.get('ความยาวตัดม้วน กะ A'),
        "ความยาวตัดม้วน กะ B (เดิม)": final.get('ความยาวตัดม้วน กะ B'),
        "น้ำหนักม้วน กะ A": np.nan,
        "น้ำหนักม้วน กะ B": np.nan,
        "จำนวนแผล กะ A": np.nan,
        "จำนวนแผล กะ B": np.nan,
        "ค่าเฉลี่ย กะ A": np.nan,
        "ค่าเฉลี่ย กะ B": np.nan,
        "Y": final.get('Y'),
        "Z": final.get('Z'),
        "ค่าแรงพนักงานกะ A": np.nan,
        "ค่าแรงพนักงานกะ B": np.nan
    })
    return export_df


def build_extra_block(df_a_raw, df_b_raw):
    extra_machine_order_old_2 = [m + " ตัดม้วนที่ครั้ง 2" for m in PRIMARY_MACHINE_ORDER]
    extra_machine_order_old_3 = [m + " ตัดม้วนที่ครั้ง 3" for m in PRIMARY_MACHINE_ORDER]
    full_search_list = (
        EXTRA_MACHINE_ORDER_FULL +
        extra_machine_order_old_2 +
        extra_machine_order_old_3
)
    df_a = df_a_raw[df_a_raw['เครื่องทอ NO'].isin(full_search_list)].copy()
    df_b = df_b_raw[df_b_raw['เครื่องทอ NO'].isin(full_search_list)].copy()
    if df_a.empty and df_b.empty:
        return pd.DataFrame()
    if not df_a.empty:
        df_a = (_ensure_cols(df_a)
                .sort_values(by='วันที่', ascending=False)
                .drop_duplicates('เครื่องทอ NO', keep='first'))
    if not df_b.empty:
        df_b = (_ensure_cols(df_b)
                .sort_values(by='วันที่', ascending=False)
                .drop_duplicates('เครื่องทอ NO', keep='first'))
    rows = []
    def add_row(r, shift):
        name = r['เครื่องทอ NO']
        if isinstance(name, str) and "ตัดม้วนที่ครั้ง" in name:
            name = name.replace("ตัดม้วนที่ครั้ง", "ตัดม้วนครั้งที่")
        row = {
            "เครื่อง": name,
            "พนักงานทอ กะ A": np.nan,
            "พนักงานทอ กะ B": np.nan,
            "ขนาดหน้าผ้า (รหัส)": r.get('ขนาดหน้าผ้า_รหัส'),
            "Lot.No": r.get('Lot.No'),
            "เลขที่ใบสั่งผลิต": r.get('เลขที่ใบสั่งผลิต'),
            "ยอดทอ กะ A": np.nan,
            "ยอดทอ กะ B": np.nan,
            "ความเร็วรอบ กะ A": np.nan,
            "ความเร็วรอบ กะ B": np.nan,
            "ประสิทธิภาพ (กะ A)%": np.nan,
            "ประสิทธิภาพ (กะ B)%": np.nan,
            "มิเตอร์เริ่มงาน กะ A": np.nan,
            "มิเตอร์เลิกงาน กะ A": np.nan,
            "มิเตอร์เริ่มงาน กะ B": np.nan,
            "มิเตอร์เลิกงาน กะ B": np.nan,
            "ความยาวตัดม้วน กะ A (เดิม)": np.nan,
            "ความยาวตัดม้วน กะ B (เดิม)": np.nan,
            "น้ำหนักม้วน กะ A": np.nan,
            "น้ำหนักม้วน กะ B": np.nan,
            "จำนวนแผล กะ A": np.nan,
            "จำนวนแผล กะ B": np.nan,
            "ค่าเฉลี่ย กะ A": np.nan,
            "ค่าเฉลี่ย กะ B": np.nan,
            "Y": np.nan,
            "Z": np.nan,
            "ค่าแรงพนักงานกะ A": np.nan,
            "ค่าแรงพนักงานกะ B": np.nan,
        }
        if shift == 'A':
            row.update({
                "พนักงานทอ กะ A": r.get('พนักงานทอ กะ A'),
                "ความเร็วรอบ กะ A": r.get('ความเร็วรอบ กะ A'),
                "ประสิทธิภาพ (กะ A)%": r.get('ประสิทธิภาพ กะ A'),
                "มิเตอร์เริ่มงาน กะ A": r.get('มิเตอร์เริ่มงาน กะ A'),
                "มิเตอร์เลิกงาน กะ A": r.get('มิเตอร์เลิกงาน กะ A'),
                "ความยาวตัดม้วน กะ A (เดิม)": r.get('ความยาวตัดม้วน กะ A'),
            })
        else:
            row.update({
                "พนักงานทอ กะ B": r.get('พนักงานทอ กะ B'),
                "ความเร็วรอบ กะ B": r.get('ความเร็วรอบ กะ B'),
                "ประสิทธิภาพ (กะ B)%": r.get('ประสิทธิภาพ กะ B'),
                "มิเตอร์เริ่มงาน กะ B": r.get('มิเตอร์เริ่มงาน กะ B'),
                "มิเตอร์เลิกงาน กะ B": r.get('มิเตอร์เลิกงาน กะ B'),
                "ความยาวตัดม้วน กะ B (เดิม)": r.get('ความยาวตัดม้วน กะ B'),
            })
        row['_sort'] = (
            extract_machine_number(name),
            detect_cut_iteration(name),
            0 if shift == 'A' else 1,
        )
        rows.append(row)
    if not df_a.empty:
        for _, r in df_a.iterrows():
            add_row(r, 'A')
    if not df_b.empty:
        for _, r in df_b.iterrows():
            add_row(r, 'B')
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df.sort_values('_sort', inplace=True)
    df.drop(columns=['_sort'], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def inject_formulas(df, excel_start_row):
    formulas_g = []
    formulas_h = []
    formulas_aa = []
    formulas_ab = []
    for i in range(len(df)):
        r = excel_start_row + i
        f_g = (
            f'=MAX(0, IF(M{r}>N{r}, '
            f'IF(Q{r}<>"", Q{r}-ABS(N{r}-M{r}), R{r}-ABS(N{r}-M{r})), '
            f'N{r}-M{r}))'
        )
        f_h = (
            f'=MAX(0, IF(O{r}>P{r}, '
            f'IF(Q{r}<>"", Q{r}-ABS(P{r}-O{r}), R{r}-ABS(P{r}-O{r})), '
            f'P{r}-O{r}))'
        )
        num_expr = (
            f'IF(LEFT(UPPER(D{r}),3)="FCL", VALUE(MID(D{r},6,2)), '
            f'IF(LEFT(UPPER(D{r}),4)="TEST", VALUE(MID(D{r},5,2)), VALUE(LEFT(D{r},2))))'
        )
        rate_expr = (
            f'IFS('
            f'OR({num_expr}=12,{num_expr}=13,{num_expr}=14,'
            f'{num_expr}=15,{num_expr}=16,{num_expr}=17),0.13,'
            f'OR({num_expr}=18,{num_expr}=19,{num_expr}=20,'
            f'{num_expr}=21,{num_expr}=22),0.14,'
            f'OR({num_expr}=23,{num_expr}=24,{num_expr}=25),0.15,'
            f'OR({num_expr}=26,{num_expr}=27,{num_expr}=28,'
            f'{num_expr}=29,{num_expr}=30),0.18)'
        )
        f_aa = (
            f'=IFERROR(TEXT(({rate_expr})*G{r},"0.00"),"คำนวณไม่สำเร็จ")'
        )
        f_ab = (
            f'=IFERROR(TEXT(({rate_expr})*H{r},"0.00"),"คำนวณไม่สำเร็จ")'
        )
        formulas_g.append(f_g)
        formulas_h.append(f_h)
        formulas_aa.append(f_aa)
        formulas_ab.append(f_ab)
    df['ยอดทอ กะ A'] = formulas_g
    df['ยอดทอ กะ B'] = formulas_h
    df['ค่าแรงพนักงานกะ A'] = formulas_aa
    df['ค่าแรงพนักงานกะ B'] = formulas_ab
    return df


def add_summary_and_extra_format(file_path, sheet_name, file3_present,
                                 extra_row_start=None, extra_row_end=None):
    try:
        wb = openpyxl.load_workbook(file_path)
        sh = wb[sheet_name]
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        underline = Font(underline="single")
        blue = Font(color="0000FF")
        red = Font(color="FF0000")
        dark_blue_font = Font(color="00008B")
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        gray_row = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
        last_col_letter = get_column_letter(sh.max_column)
        for row in sh[f'A2:{last_col_letter}{MAX_ROW_FOR_BORDER_PRIMARY}']:
            for c in row:
                c.border = border
                if c.row % 2 == 0:
                    c.fill = gray_row
        now = datetime.now(ZoneInfo("Asia/Bangkok"))
        th_year = now.year + 543
        date_label = now.strftime(f"%d-%m-{th_year}")
        sh.merge_cells('A88:B88')
        a88 = sh['A88']
        a88.value = f"สรุปยอดผลิตประจำวันที่ ({date_label})"
        a88.alignment = center
        a88.fill = gray
        for row in sh['A88:B88']:
            for c in row:
                c.border = border
        sh['A91'].value = "ยอดทอกะ A ทั้งหมด"
        sh['A91'].alignment = center
        sh['A91'].font = underline
        sh['B91'].value = "ยอดทอกะ B ทั้งหมด"
        sh['B91'].alignment = center
        sh['B91'].font = underline
        sh['A93'].value = f"=SUBTOTAL(109,G2:G{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['A93'].alignment = center
        sh['A93'].number_format = '#,##0'
        sh['A93'].border = border
        sh['B93'].value = f"=SUBTOTAL(109,H2:H{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['B93'].alignment = center
        sh['B93'].number_format = '#,##0'
        sh['B93'].border = border
        sh.merge_cells('A94:B94')
        sh['A94'].value = "=A93+B93"
        sh['A94'].alignment = center
        sh['A94'].number_format = '#,##0'
        for row in sh['A94:B94']:
            for c in row:
                c.border = border
        sh['A97'].value = "ยอดตัดม้วนกะ A ทั้งหมด"
        sh['A97'].alignment = center
        sh['A97'].font = underline
        sh['B97'].value = "ยอดตัดม้วนกะ B ทั้งหมด"
        sh['B97'].alignment = center
        sh['B97'].font = underline
        sh['A99'].value = f"=SUBTOTAL(109,Q2:Q{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['A99'].alignment = center
        sh['A99'].number_format = '#,##0'
        sh['A99'].border = border
        sh['B99'].value = f"=SUBTOTAL(109,R2:R{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['B99'].alignment = center
        sh['B99'].number_format = '#,##0'
        sh['B99'].border = border
        sh.merge_cells('A100:B100')
        sh['A100'].value = "=A99+B99"
        sh['A100'].alignment = center
        sh['A100'].number_format = '#,##0'
        for row in sh['A100:B100']:
            for c in row:
                c.border = border
        sh['A103'].value = f"=SUBTOTAL(109,Y2:Y{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['A103'].alignment = center
        sh['A103'].number_format = '#,##0'
        sh['A103'].border = border
        sh['B103'].value = f"=SUBTOTAL(109,Z2:Z{MAX_ROW_FOR_BORDER_PRIMARY})"
        sh['B103'].alignment = center
        sh['B103'].number_format = '#,##0'
        sh['B103'].border = border
        if file3_present:
            sh['A103'].font = blue
            sh['B103'].font = red
        sh.merge_cells('A104:B104')
        sh['A104'].value = "=A103+B103"
        sh['A104'].alignment = center
        sh['A104'].number_format = '#,##0'
        for row in sh['A104:B104']:
            for c in row:
                c.border = border
        sh.auto_filter.ref = f"A1:{last_col_letter}{MAX_ROW_FOR_BORDER_PRIMARY}"
        sh.freeze_panes = 'A2'
        if extra_row_start and extra_row_end and extra_row_end >= extra_row_start:
            for row in sh[f'A{extra_row_start}:{last_col_letter}{extra_row_end}']:
                for c in row:
                    c.border = border
                    if c.row % 2 == 0:
                        c.fill = gray_row
            cut2_rows = []
            cut3_rows = []
            for r in range(extra_row_start, extra_row_end + 1):
                v = sh[f"A{r}"].value
                if isinstance(v, str):
                    if ("ตัดม้วนที่ครั้ง 2" in v) or ("ตัดม้วนครั้งที่ 2" in v):
                        cut2_rows.append(r)
                    elif ("ตัดม้วนที่ครั้ง 3" in v) or ("ตัดม้วนครั้งที่ 3" in v):
                        cut3_rows.append(r)
            def simple_sum_formula(col, rows):
                if not rows:
                    return "0"
                parts = [f"{col}{r}" for r in rows]
                return f"=SUM({','.join(parts)})"
            c2_g = simple_sum_formula("G", cut2_rows)
            c2_h = simple_sum_formula("H", cut2_rows)
            c2_q = simple_sum_formula("Q", cut2_rows)
            c2_r = simple_sum_formula("R", cut2_rows)
            c3_g = simple_sum_formula("G", cut3_rows)
            c3_h = simple_sum_formula("H", cut3_rows)
            c3_q = simple_sum_formula("Q", cut3_rows)
            c3_r = simple_sum_formula("R", cut3_rows)
            sh['I91'].value = "ยอดทอกะ A (ตัดครั้ง 2)"
            sh['I91'].alignment = center
            sh['I91'].font = underline
            sh['J91'].value = "ยอดทอกะ B (ตัดครั้ง 2)"
            sh['J91'].alignment = center
            sh['J91'].font = underline
            sh['I93'].value = c2_g
            sh['I93'].alignment = center
            sh['I93'].number_format = '#,##0'
            sh['I93'].border = border
            sh['I93'].font = dark_blue_font
            sh['J93'].value = c2_h
            sh['J93'].alignment = center
            sh['J93'].number_format = '#,##0'
            sh['J93'].border = border
            sh['J93'].font = dark_blue_font
            sh.merge_cells('I94:J94')
            sh['I94'].value = "=I93+J93"
            sh['I94'].alignment = center
            sh['I94'].number_format = '#,##0'
            sh['I94'].font = dark_blue_font
            for row in sh['I94:J94']:
                for c in row:
                    c.border = border
            sh['I97'].value = "ยอดตัดม้วนกะ A (ครั้ง 2)"
            sh['I97'].alignment = center
            sh['I97'].font = underline
            sh['J97'].value = "ยอดตัดม้วนกะ B (ครั้ง 2)"
            sh['J97'].alignment = center
            sh['J97'].font = underline
            sh['I99'].value = c2_q
            sh['I99'].alignment = center
            sh['I99'].number_format = '#,##0'
            sh['I99'].border = border
            sh['I99'].font = dark_blue_font
            sh['J99'].value = c2_r
            sh['J99'].alignment = center
            sh['J99'].number_format = '#,##0'
            sh['J99'].border = border
            sh['J99'].font = dark_blue_font
            sh.merge_cells('I100:J100')
            sh['I100'].value = "=I99+J99"
            sh['I100'].alignment = center
            sh['I100'].number_format = '#,##0'
            sh['I100'].font = dark_blue_font
            for row in sh['I100:J100']:
                for c in row:
                    c.border = border
            sh['L91'].value = "ยอดทอกะ A (ตัดครั้ง 3)"
            sh['L91'].alignment = center
            sh['L91'].font = underline
            sh['M91'].value = "ยอดทอกะ B (ตัดครั้ง 3)"
            sh['M91'].alignment = center
            sh['M91'].font = underline
            sh['L93'].value = c3_g
            sh['L93'].alignment = center
            sh['L93'].number_format = '#,##0'
            sh['L93'].border = border
            sh['L93'].font = dark_blue_font
            sh['M93'].value = c3_h
            sh['M93'].alignment = center
            sh['M93'].number_format = '#,##0'
            sh['M93'].border = border
            sh['M93'].font = dark_blue_font
            sh.merge_cells('L94:M94')
            sh['L94'].value = "=L93+M93"
            sh['L94'].alignment = center
            sh['L94'].number_format = '#,##0'
            sh['L94'].font = dark_blue_font
            for row in sh['L94:M94']:
                for c in row:
                    c.border = border
            sh['L97'].value = "ยอดตัดม้วนกะ A (ครั้ง 3)"
            sh['L97'].alignment = center
            sh['L97'].font = underline
            sh['M97'].value = "ยอดตัดม้วนกะ B (ครั้ง 3)"
            sh['M97'].alignment = center
            sh['M97'].font = underline
            sh['L99'].value = c3_q
            sh['L99'].alignment = center
            sh['L99'].number_format = '#,##0'
            sh['L99'].border = border
            sh['L99'].font = dark_blue_font
            sh['M99'].value = c3_r
            sh['M99'].alignment = center
            sh['M99'].number_format = '#,##0'
            sh['M99'].border = border
            sh['M99'].font = dark_blue_font
            sh.merge_cells('L100:M100')
            sh['L100'].value = "=L99+M99"
            sh['L100'].alignment = center
            sh['L100'].number_format = '#,##0'
            sh['L100'].font = dark_blue_font
            for row in sh['L100:M100']:
                for c in row:
                    c.border = border
        wb.save(file_path)
    except Exception as e:
        print(f"[เกิดข้อผิดพลาดส่วนสรุป] {e}")


def calc_display_width(text):
    if text is None:
        return 0
    s = str(text)
    if s == 'nan':
        return 0
    wide = sum(1 for ch in s if ord(ch) > 255)
    narrow = len(s) - wide
    return narrow + wide * 1.15

def build_column_width_map(primary_df, extra_df=None, header_list=None,
                            min_w=6, max_w=48, formula_cols=None):
    if header_list is None:
        header_list = list(primary_df.columns)
    if formula_cols is None:
        formula_cols = {
            "ยอดทอ กะ A","ยอดทอ กะ B","ค่าแรงพนักงานกะ A","ค่าแรงพนักงานกะ B"
        }
    width_map = {}
    dfs = [primary_df]
    if extra_df is not None and not extra_df.empty:
        dfs.append(extra_df)
    for col in header_list:
        max_width = calc_display_width(col)
        for d in dfs:
            if col in d.columns:
                if col == 'Lot.No':
                    max_width = max(
                        max_width,
                        calc_display_width("68091601") + 2
)
                    continue
                for val in d[col].head(500):
                    if col in formula_cols:
                        max_width = max(max_width, 12)
                    else:
                        max_width = max(
                            max_width,
                            calc_display_width(val)
)
        max_width += 2
        max_width = max(min_w, min(max_w, max_width))
        width_map[col] = max_width
    return width_map


print("\n[ขั้นที่ 1/3] เลือกไฟล์กะ A (report A)")
up_a = files.upload()
if not up_a:
    raise ValueError("ไม่พบไฟล์กะ A")
file_a_name = next(iter(up_a))
file_a_content = up_a[file_a_name]

print("\n[ขั้นที่ 2/3] เลือกไฟล์กะ B (report B)")
up_b = files.upload()
if not up_b:
    raise ValueError("ไม่พบไฟล์กะ B")
file_b_name = next(iter(up_b))
file_b_content = up_b[file_b_name]
print("\n[ขั้นที่ 3/3] (ถ้ามี) เลือกไฟล์ตัดม้วนเสริม (ไฟล์ C)")
up_c = None
try:
    up_c = files.upload()
    if up_c:
        file_c_name = next(iter(up_c))
        file_c_content = up_c[file_c_name]
        print(f"อ่านไฟล์ C: {file_c_name}")
    else:
        print("ข้ามไฟล์ C")
except Exception:
    print("ข้ามไฟล์ C")
df_a_raw = read_generic_file(
    file_a_content, file_a_name, EXPECTED_COLUMN_COUNT_AB, RAW_COLUMNS_AB
)
df_b_raw = read_generic_file(
    file_b_content, file_b_name, EXPECTED_COLUMN_COUNT_AB, RAW_COLUMNS_AB
)
df_c_raw = read_generic_file(
    file_c_content, file_c_name, EXPECTED_COLUMN_COUNT_C, RAW_COLUMNS_C
) if up_c else pd.DataFrame()
FILE3_PRESENT = not df_c_raw.empty
if df_a_raw.empty and df_b_raw.empty:
    raise ValueError("ไม่มีข้อมูลทั้งกะ A และ B")
primary_df = build_primary_block(df_a_raw, df_b_raw, df_c_raw)
extra_df = build_extra_block(df_a_raw, df_b_raw)
has_extra = not extra_df.empty
primary_df = assign_production_code(primary_df)
if has_extra:
    extra_df = assign_production_code(extra_df)
if has_extra:
    combined_df = pd.concat([primary_df, extra_df], ignore_index=True)
else:
    combined_df = primary_df.copy()
combined_df = generate_and_assign_lot_no(combined_df)
primary_len = len(primary_df)
primary_df = combined_df.iloc[:primary_len].copy()
if has_extra:
    extra_df = combined_df.iloc[primary_len:].copy()
primary_df = inject_formulas(primary_df, excel_start_row=2)
for col in NUMERIC_DATA_COLUMNS:
    if col in primary_df.columns and col not in ["ยอดทอ กะ A","ยอดทอ กะ B"]:
        primary_df[col] = pd.to_numeric(primary_df[col], errors='coerce')
if has_extra:
    extra_df = inject_formulas(
        extra_df, excel_start_row=EXTRA_BLOCK_DATA_START_EXCEL_ROW
)
    for col in NUMERIC_DATA_COLUMNS:
        if col in extra_df.columns and col not in ["ยอดทอ กะ A","ยอดทอ กะ B"]:
            extra_df[col] = pd.to_numeric(extra_df[col], errors='coerce')
header_display = list(primary_df.columns)
if 'ความยาวตัดม้วน กะ A (เดิม)' in header_display:
    idx_a = header_display.index('ความยาวตัดม้วน กะ A (เดิม)')
    header_display[idx_a] = 'ความยาวตัดม้วน กะ A'
if 'ความยาวตัดม้วน กะ B (เดิม)' in header_display:
    idx_b = header_display.index('ความยาวตัดม้วน กะ B (เดิม)')
    header_display[idx_b] = 'ความยาวตัดม้วน กะ B'
if 'Y' in header_display:
    header_display[header_display.index('Y')] = 'ความยาวตัดม้วน กะ A'
if 'Z' in header_display:
    header_display[header_display.index('Z')] = 'ความยาวตัดม้วน กะ B'
width_map = build_column_width_map(
    primary_df, extra_df if has_extra else None, header_display
)
now = datetime.now(ZoneInfo('Asia/Bangkok'))
thai_year = now.year + 543
timestamp = now.strftime(f"%d-%m-{thai_year}_%H%M%S")
output_filename = f"รายงานสรุปผลผลิตรายเครื่อง_{timestamp}.xlsx"
writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
wb = writer.book
ws_name = 'รายงานรวม'
fmt_header = wb.add_format({
    'bold': True,
    'text_wrap': True,
    'valign': 'vcenter',
    'align': 'center',
    'fg_color': '#DDEBF7',
    'border': 1,
})
fmt_text = wb.add_format({'align': 'left', 'valign': 'vcenter'})
fmt_num_black = wb.add_format({'align': 'right', 'valign': 'vcenter', 'num_format': '#,##0'})
fmt_blue = wb.add_format({
    'align': 'right','valign': 'vcenter','num_format': '#,##0','font_color': 'blue',
})
fmt_red = wb.add_format({
    'align': 'right','valign': 'vcenter','num_format': '#,##0','font_color': 'red',
})
primary_df.to_excel(
    writer, sheet_name=ws_name, startrow=1, header=False, index=False
)
ws = writer.sheets[ws_name]
ws.write_row(0, 0, header_display, fmt_header)
display_to_internal = {
    disp: internal for disp, internal in zip(header_display, list(primary_df.columns))
}
for col_idx, disp_name in enumerate(header_display):
    internal_name = display_to_internal[disp_name]
    width = width_map.get(disp_name, 14)
    if internal_name in NUMERIC_DATA_COLUMNS:
        if internal_name == 'Y':
            ws.set_column(
                col_idx, col_idx, width,
                fmt_blue if FILE3_PRESENT else fmt_num_black
)
        elif internal_name == 'Z':
            ws.set_column(
                col_idx, col_idx, width,
                fmt_red if FILE3_PRESENT else fmt_num_black
)
        else:
            ws.set_column(col_idx, col_idx, width, fmt_num_black)
    else:
        ws.set_column(col_idx, col_idx, width, fmt_text)
ws.set_column(16, 16, width_map.get(header_display[16], 14), fmt_num_black)
ws.set_column(17, 17, width_map.get(header_display[17], 14), fmt_num_black)
if has_extra:
    start_data_zero = EXTRA_BLOCK_DATA_START_EXCEL_ROW - 1
    header_extra_zero = EXTRA_BLOCK_HEADER_EXCEL_ROW - 1
    extra_df.to_excel(
        writer, sheet_name=ws_name, startrow=start_data_zero,
        header=False, index=False
)
    ws.write_row(header_extra_zero, 0, header_display, fmt_header)
    ws.set_column(16, 16, width_map.get(header_display[16], 14), fmt_num_black)
    ws.set_column(17, 17, width_map.get(header_display[17], 14), fmt_num_black)
else:
    print("ไม่มีข้อมูลตัดม้วนเพิ่ม (ครั้ง 2/3)")
writer.close()
extra_start = EXTRA_BLOCK_DATA_START_EXCEL_ROW if has_extra else None
extra_end = (
    EXTRA_BLOCK_DATA_START_EXCEL_ROW + len(extra_df) - 1 if has_extra else None
)
add_summary_and_extra_format(
    output_filename, ws_name, FILE3_PRESENT,
    extra_row_start=extra_start, extra_row_end=extra_end
)
add_size_code_summaries = None
def add_size_code_summaries(file_path, sheet_name, start_row=150):
    try:
        wb = openpyxl.load_workbook(file_path)
        sh = wb[sheet_name]
        last_data_row = sh.max_row
        if extra_end:
            last_data_row = extra_end
        elif primary_df is not None and not primary_df.empty:
            last_data_row = min(MAX_ROW_FOR_BORDER_PRIMARY, 1 + len(primary_df))
        codes = []
        for r in range(2, last_data_row + 1):
            if r == 109:
                continue
            cell_value = sh.cell(row=r, column=4).value
            if cell_value:
                code_str = str(cell_value).strip()
                if code_str and code_str not in codes:
                    codes.append(code_str)
        if not codes:
            wb.save(file_path)
            return
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_r = start_row - 1
        sh.cell(row=header_r, column=1).value = "สรุปผลผลิตตามรหัสผ้า"
        sh.cell(row=header_r, column=2).value = "สรุปยอดทอ"
        sh.cell(row=header_r, column=3).value = "สรุปยอดตัดม้วน"
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True)
        for c in range(1, 4):
            cell = sh.cell(row=header_r, column=c)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
        try:
            b_header = "Total produced"
            c_header = "Total cut length"
            w_b = calc_display_width(b_header) + 4
            w_c = calc_display_width(c_header) + 4
            sh.column_dimensions[get_column_letter(2)].width = w_b
            sh.column_dimensions[get_column_letter(3)].width = w_c
        except Exception:
            sh.column_dimensions[get_column_letter(2)].width = 20
            sh.column_dimensions[get_column_letter(3)].width = 32
        for idx, code in enumerate(codes):
            r = start_row + idx
            sh.cell(row=r, column=1).value = code
            sh.cell(row=r, column=2).value = (
                f'=SUMIF($D$2:$D${last_data_row}, $A${r}, $G$2:$G${last_data_row}) + '
                f'SUMIF($D$2:$D${last_data_row}, $A${r}, $H$2:$H${last_data_row})'
)
            sh.cell(row=r, column=3).value = (
                f'=SUMIF($D$2:$D${last_data_row}, $A${r}, $Q$2:$Q${last_data_row}) + '
                f'SUMIF($D$2:$D${last_data_row}, $A${r}, $R$2:$R${last_data_row})'
)
            col_a_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            col_b_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            col_c_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for c in range(1, 4):
                cell = sh.cell(row=r, column=c)
                cell.border = border
                if c == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = col_a_fill
                elif c == 2:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '#,##0'
                    cell.fill = col_b_fill
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '#,##0'
                    cell.fill = col_c_fill
        wb.save(file_path)
    except Exception as e:
        print(f"[ข้อผิดพลาด add_size_code_summaries] {e}")
add_size_code_summaries(output_filename, ws_name, start_row=150)
print("-" * 72)
print(f"สร้างรายงานเสร็จ: {output_filename}")
print(f"มีไฟล์ C: {'Yes' if FILE3_PRESENT else 'No'}")
files.download(output_filename)